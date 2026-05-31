# -*- coding: utf-8 -*-
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_csv_to_beautiful_xlsx():
    # 스크립트 파일이 위치한 디렉토리로 작업 디렉토리 변경
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if script_dir:
        os.chdir(script_dir)
        
    csv_file = "yanggu_all_crop_transitions.csv"
    xlsx_file = "yanggu_all_crop_transitions.xlsx"
    
    if not os.path.exists(csv_file):
        print(f"[오류] 원본 CSV 파일이 존재하지 않습니다: {csv_file}")
        return
        
    wb = Workbook()
    ws = wb.active
    ws.title = "양구군 경작지 전환 대조표"
    
    # 그리드라인 보이기 설정
    ws.views.sheetView[0].showGridLines = True
    
    # 폰트 설정 (맑은 고딕 또는 프리미엄 폰트 대체용)
    font_title = Font(name="맑은 고딕", size=16, bold=True, color="1F4E79")
    font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="맑은 고딕", size=10)
    font_note = Font(name="맑은 고딕", size=9, italic=True, color="555555")
    
    # 채우기 색상 설정 (Hex)
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    
    # 테두리 설정
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # 정렬 설정
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 1. 상단 타이틀 추가
    ws.merge_cells("A1:L1")
    ws["A1"] = "양구군 농경지 경작물 대체 실태 대조 분석표 (2022~2026)"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 2. 설명 문구 추가
    ws.merge_cells("A2:L2")
    ws["A2"] = "* 분석 기준 데이터: 미국 USGS/Copernicus Sentinel-2 및 Landsat 시계열 식생 지수(NDVI)와 고해상도 하이브리드 정사영상 대조 자료"
    ws["A2"].font = font_note
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # 3. CSV 데이터 로드 및 시트 작성
    with open(csv_file, mode='r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers = next(reader)
        
        # 헤더 행 작성
        header_row = 4
        ws.row_dimensions[header_row].height = 28
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_data
            
        # 데이터 행 작성
        data_start_row = 5
        for row_idx, row_data in enumerate(reader, data_start_row):
            ws.row_dimensions[row_idx].height = 24
            is_zebra = (row_idx % 2 == 0)
            
            for col_idx, val in enumerate(row_data, 1):
                # 숫자 변환 시도
                if col_idx in [1, 9, 10]: # 일련번호, 위도, 경도
                    try:
                        val = float(val) if '.' in val else int(val)
                    except ValueError:
                        pass
                        
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_data
                cell.border = border_data
                
                # 제브라 패턴 채우기
                if is_zebra:
                    cell.fill = fill_zebra
                    
                # 컬럼 특성에 따른 정렬
                if col_idx in [1, 2, 4, 5, 6, 7, 8, 11, 12]:  # ID, 읍면, 지목, 작물명, 대체여부, 확인일 등
                    cell.alignment = align_center
                else:  # 소재지 주소, 위성 판독 근거
                    cell.alignment = align_left
                    
    # 4. 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            # 병합된 타이틀 행은 너비 계산에서 제외
            if cell.row in [1, 2]:
                continue
            if cell.value:
                # 한글 문자 길이 보정을 위한 길이 계산
                val_str = str(cell.value)
                byte_len = len(val_str.encode('utf-8'))
                # 바이트 길이를 대략적인 글자 너비로 환산
                max_len = max(max_len, byte_len)
                
        # 적정 너비 산정 (최대 60 글자 너비로 제한하되 줄바꿈 지원)
        adjusted_width = max(max_len // 2 + 3, 10)
        if adjusted_width > 60:
            adjusted_width = 60
            
        # 주소와 위성 판독 근거는 넉넉하게 크기 고정
        if col_letter == 'C': # 소재지 주소
            adjusted_width = 38
        elif col_letter == 'I': # 위성 판독 근거
            adjusted_width = 50
            
        ws.column_dimensions[col_letter].width = adjusted_width
        
    wb.save(xlsx_file)
    print(f"[성공] 고품격 엑셀 파일 변환이 완료되었습니다: {xlsx_file}")
    print(f"경로: {os.path.abspath(xlsx_file)}")

if __name__ == "__main__":
    convert_csv_to_beautiful_xlsx()
