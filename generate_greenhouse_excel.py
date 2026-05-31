# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_greenhouse_excel():
    xlsx_file = "greenhouse_parts_list.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "육각 목재 비닐하우스 부품"
    
    # 그리드라인 보이기 설정
    ws.views.sheetView[0].showGridLines = True
    
    # 폰트 설정 (맑은 고딕 또는 프리미엄 폰트 대체용)
    font_title = Font(name="맑은 고딕", size=16, bold=True, color="2C5E3B") # Forest Green
    font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="맑은 고딕", size=10)
    font_formula = Font(name="맑은 고딕", size=10, bold=True, color="000000")
    font_note = Font(name="맑은 고딕", size=9, italic=True, color="555555")
    
    # 채우기 색상 설정 (Hex)
    fill_header = PatternFill(start_color="2C5E3B", end_color="2C5E3B", fill_type="solid") # Dark Forest Green
    fill_zebra = PatternFill(start_color="F2F7F3", end_color="F2F7F3", fill_type="solid")  # Pale Sage Green
    fill_total = PatternFill(start_color="E6EFEA", end_color="E6EFEA", fill_type="solid")  # Soft Green Tint
    
    # 테두리 설정
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_side = Side(style='double', color='000000')
    border_total = Border(top=thin_border_side, bottom=double_bottom_side, left=thin_border_side, right=thin_border_side)
    
    # 정렬 설정
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # 1. 상단 타이틀 추가
    ws.merge_cells("A1:I1")
    ws["A1"] = "육각형 목재 비닐하우스 정밀 부품 목록 및 견적 (Hexagonal Wooden Greenhouse)"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 2. 설명 문구 추가
    ws.merge_cells("A2:I2")
    ws["A2"] = "* 구조 규격: 외경 6.0m (반경 3.0m), 지상 기둥 높이 2.4m, 육각 지붕 정점 높이 3.2m / 주요 자재: 미송 방부목(H3 등급) 및 0.15mm PO 방무 비닐 피복"
    ws["A2"].font = font_note
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # 3. 헤더 정의
    headers = ["순번", "자재 분류", "부품명", "상세 규격(mm/t/외)", "수량", "단위", "재질", "예상 단가(KRW)", "예상 금액(KRW)"]
    header_row = 4
    ws.row_dimensions[header_row].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_data
        
    # 4. 데이터 로드
    raw_data = [
        # 구조재 (Lumber)
        ["구조재", "메인 기둥 (Post)", "90 x 90 x 2,400", 6, "EA", "낙엽송 방부목 (H3)", 45000, "구조 버팀 기둥"],
        ["구조재", "하부 지름 링 가로대", "38 x 140 x 3,000 (2x6)", 6, "EA", "미송 방부목 (H3)", 24000, "하단 기단 프레임"],
        ["구조재", "상부 헤더 가로대", "38 x 140 x 3,000 (2x6)", 6, "EA", "SPF 구조재", 22000, "상단 벽체 고정 가로대"],
        ["구조재", "지붕 서까래 (Rafter)", "38 x 89 x 3,500 (2x4)", 6, "EA", "SPF 구조재", 18000, "육각 지붕 골조"],
        ["구조재", "도어 입구 프레임", "38 x 89 x 2,100 (2x4)", 4, "EA", "SPF 구조재", 14000, "출입구 보강 프레임"],
        ["구조재", "창문 환기구 프레임", "38 x 38 x 1,200 (2x2)", 12, "EA", "SPF 구조재", 6000, "측면 개폐창 뼈대"],
        # 조립 부속 (Hardware)
        ["조립 부속", "지붕 센트럴 헥사 허브", "6.0t 두께, 아연도금 마감", 1, "EA", "SS400 스틸", 120000, "지붕 중심 연결 철물"],
        ["조립 부속", "포스트 메탈 베이스 (U자)", "90x90용 하부 고정쇠", 6, "EA", "SUS304 스테인리스", 25000, "콘크리트 기단 앵커용"],
        ["조립 부속", "L형 모서리 보강 브라켓", "80 x 80 x 4.0t", 24, "EA", "SUS304 스테인리스", 3500, "벽체-기둥 접합부 보강"],
        ["조립 부속", "고장력 델타 목재 피스", "직경 5.0 x 길이 75", 1200, "EA", "아연도금강", 50, "골조 결합 나사못"],
        ["조립 부속", "셋트 앵커 볼트", "M12 x 100", 12, "EA", "SUS304 스테인리스", 2500, "기단 베이스 고정용"],
        ["조립 부속", "도어용 힌지 및 시클 잠금장치", "헤비 듀티 타입 락 세트", 1, "Set", "아연도금강", 35000, "출입문 힌지 및 개폐 잠금쇠"],
        # 피복 및 외장 (Covering)
        ["피복 및 외장", "PO 하우스 광학 비닐", "두께 0.15mm / 폭 10m", 30, "m", "PO 방무적 3년필름", 15000, "전체 외장 피복 필름"],
        ["피복 및 외장", "비닐 고정용 사각 패드", "두께 0.8t / 폭 30 / 길이 3,000", 16, "EA", "알루미늄 프로파일", 8000, "비닐 압착용 채널 레일"],
        ["피복 및 외장", "패드 고정용 복원 스프링", "길이 2,000 (갈지자 형)", 160, "EA", "아연도금 스프링강", 800, "패드 내 비닐 압착선"],
        ["피복 및 외장", "방충망 내장 수동 개폐기", "롤업 환기 핸들 세트", 2, "Set", "복합재료 및 기어", 65000, "측면 자연 환기 개폐"],
        ["피복 및 외장", "이중 부틸 방수 테이프", "폭 50mm x 길이 15m", 5, "Roll", "부틸 고무", 12000, "비닐 겹침 및 프레임 씰링"]
    ]
    
    start_row = 5
    for i, row_data in enumerate(raw_data, start_row):
        ws.row_dimensions[i].height = 24
        is_zebra = (i % 2 == 0)
        
        # 순번
        cell_num = ws.cell(row=i, column=1, value=i - start_row + 1)
        cell_num.font = font_data
        cell_num.alignment = align_center
        cell_num.border = border_data
        
        # 분류
        cell_cat = ws.cell(row=i, column=2, value=row_data[0])
        cell_cat.font = font_data
        cell_cat.alignment = align_center
        cell_cat.border = border_data
        
        # 부품명
        cell_name = ws.cell(row=i, column=3, value=row_data[1])
        cell_name.font = font_data
        cell_name.alignment = align_left
        cell_name.border = border_data
        
        # 상세규격
        cell_spec = ws.cell(row=i, column=4, value=row_data[2])
        cell_spec.font = font_data
        cell_spec.alignment = align_center
        cell_spec.border = border_data
        
        # 수량
        cell_qty = ws.cell(row=i, column=5, value=row_data[3])
        cell_qty.font = font_data
        cell_qty.alignment = align_center
        cell_qty.border = border_data
        
        # 단위
        cell_unit = ws.cell(row=i, column=6, value=row_data[4])
        cell_unit.font = font_data
        cell_unit.alignment = align_center
        cell_unit.border = border_data
        
        # 재질
        cell_mat = ws.cell(row=i, column=7, value=row_data[5])
        cell_mat.font = font_data
        cell_mat.alignment = align_center
        cell_mat.border = border_data
        
        # 단가 (숫자)
        cell_price = ws.cell(row=i, column=8, value=row_data[6])
        cell_price.font = font_data
        cell_price.number_format = '₩#,##0'
        cell_price.alignment = align_right
        cell_price.border = border_data
        
        # 금액 (수식: 수량 * 단가)
        # 엑셀 수식 '=E{row}*H{row}' 형태로 적용하여 엑셀 프로그램에서 자동 연산되도록 설계
        formula = f"=E{i}*H{i}"
        cell_total = ws.cell(row=i, column=9, value=formula)
        cell_total.font = font_data
        cell_total.number_format = '₩#,##0'
        cell_total.alignment = align_right
        cell_total.border = border_data
        
        if is_zebra:
            for col in range(1, 10):
                ws.cell(row=i, column=col).fill = fill_zebra
                
    # 5. 합계 행 추가
    total_row = start_row + len(raw_data)
    ws.row_dimensions[total_row].height = 28
    
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    cell_lbl = ws.cell(row=total_row, column=1, value="총 자재 공급 합계 (Supply Total)")
    cell_lbl.font = Font(name="맑은 고딕", size=11, bold=True)
    cell_lbl.alignment = align_center
    cell_lbl.fill = fill_total
    
    # 합계 테두리 채우기 (병합된 셀 영역 테두리 그리기 위해 루프 사용)
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).border = border_total
        ws.cell(row=total_row, column=col).fill = fill_total
        
    # 단가 열 빈칸 마감
    cell_empty = ws.cell(row=total_row, column=8, value="")
    cell_empty.border = border_total
    cell_empty.fill = fill_total
    
    # 금액 합계 수식 (=SUM(I5:I{total_row-1}))
    sum_formula = f"=SUM(I{start_row}:I{total_row-1})"
    cell_sum = ws.cell(row=total_row, column=9, value=sum_formula)
    cell_sum.font = font_formula
    cell_sum.number_format = '₩#,##0'
    cell_sum.alignment = align_right
    cell_sum.fill = fill_total
    cell_sum.border = border_total
    
    # 부가세 및 종합 행 추가
    vat_row = total_row + 1
    ws.row_dimensions[vat_row].height = 24
    ws.merge_cells(start_row=vat_row, start_column=1, end_row=vat_row, end_column=8)
    ws.cell(row=vat_row, column=1, value="부가가치세 (VAT 10%)").font = font_data
    ws.cell(row=vat_row, column=1).alignment = align_right
    ws.cell(row=vat_row, column=9, value=f"=I{total_row}*0.1").font = font_data
    ws.cell(row=vat_row, column=9).number_format = '₩#,##0'
    ws.cell(row=vat_row, column=9).alignment = align_right
    for col in range(1, 10):
        ws.cell(row=vat_row, column=col).border = border_data
        
    g_total_row = vat_row + 1
    ws.row_dimensions[g_total_row].height = 28
    ws.merge_cells(start_row=g_total_row, start_column=1, end_row=g_total_row, end_column=8)
    ws.cell(row=g_total_row, column=1, value="총 구입 예산 합계 (Grand Total)").font = Font(name="맑은 고딕", size=11, bold=True, color="8B0000")
    ws.cell(row=g_total_row, column=1).alignment = align_right
    ws.cell(row=g_total_row, column=9, value=f"=I{total_row}+I{vat_row}").font = Font(name="맑은 고딕", size=11, bold=True, color="8B0000")
    ws.cell(row=g_total_row, column=9).number_format = '₩#,##0'
    ws.cell(row=g_total_row, column=9).alignment = align_right
    for col in range(1, 10):
        ws.cell(row=g_total_row, column=col).border = border_total
        ws.cell(row=g_total_row, column=col).fill = fill_total

    # 6. 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # 타이틀과 긴 주석행 제외
            if cell.row in [1, 2, total_row, vat_row, g_total_row]:
                continue
            if cell.value:
                val_str = str(cell.value)
                # 한글 문자 가중치 부여 길이
                byte_len = len(val_str.encode('utf-8'))
                max_len = max(max_len, byte_len)
        
        adjusted_width = max(max_len // 2 + 3, 10)
        # 특정 컬럼 조정 고정
        if col_letter == 'C': # 부품명
            adjusted_width = 30
        elif col_letter == 'D': # 상세규격
            adjusted_width = 25
        elif col_letter == 'G': # 재질
            adjusted_width = 22
            
        ws.column_dimensions[col_letter].width = adjusted_width
        
    wb.save(xlsx_file)
    print(f"[성공] 엑셀 자재 목록이 생성되었습니다: {xlsx_file}")

if __name__ == "__main__":
    generate_greenhouse_excel()
